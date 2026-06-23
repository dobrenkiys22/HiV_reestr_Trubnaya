import os
import json
import re
import base64
import time
import threading
import uuid
from io import BytesIO

import openpyxl
from flask import Flask, request
import requests

app = Flask(__name__)

TELEGRAM_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
APPS_SCRIPT_URL = os.environ["APPS_SCRIPT_URL"]
MANAGER_CHAT_ID = os.environ["MANAGER_CHAT_ID"]
SHEET_NAME = os.environ.get("SHEET_NAME", "июнь 2026")

TELEGRAM_API = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"

# Временное хранилище альбомов (несколько фото одной накладной, отправленных как альбом)
pending_albums = {}
ALBUM_FLUSH_DELAY = 3  # секунд ожидания, прежде чем считать альбом собранным
albums_lock = threading.Lock()

# Накладные, ожидающие подтверждения суммы от менеджера (когда ИИ не уверен)
pending_confirmations = {}
confirmations_lock = threading.Lock()

# Чаты, где менеджер должен прислать сумму вручную текстом (после нажатия "Другое")
awaiting_manual_sum = {}
awaiting_lock = threading.Lock()


def tg_get_file_path(file_id):
    r = requests.get(f"{TELEGRAM_API}/getFile", params={"file_id": file_id}, timeout=30)
    r.raise_for_status()
    return r.json()["result"]["file_path"]


def tg_download_file(file_path):
    r = requests.get(f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}", timeout=30)
    r.raise_for_status()
    return r.content


def tg_send_message(chat_id, text, reply_markup=None):
    payload = {"chat_id": chat_id, "text": text}
    if reply_markup:
        payload["reply_markup"] = json.dumps(reply_markup)
    try:
        r = requests.post(f"{TELEGRAM_API}/sendMessage", json=payload, timeout=15)
        return r.json()
    except Exception as e:
        print("Не удалось отправить сообщение в Telegram:", e)
        return None


def tg_answer_callback_query(callback_query_id, text=None):
    try:
        payload = {"callback_query_id": callback_query_id}
        if text:
            payload["text"] = text
        requests.post(f"{TELEGRAM_API}/answerCallbackQuery", json=payload, timeout=15)
    except Exception as e:
        print("Не удалось ответить на callback_query:", e)


def tg_remove_keyboard(chat_id, message_id):
    try:
        requests.post(
            f"{TELEGRAM_API}/editMessageReplyMarkup",
            json={"chat_id": chat_id, "message_id": message_id, "reply_markup": json.dumps({"inline_keyboard": []})},
            timeout=15,
        )
    except Exception as e:
        print("Не удалось убрать кнопки:", e)


SYSTEM_PROMPT = (
    "Ты распознаёшь фото бумажных приходных накладных ресторана "
    "(УПД / счёт-фактура / товарно-транспортная накладная). "
    "Тебе может быть передано несколько фото одного документа (например, разные страницы "
    "или продолжение таблицы) — анализируй их как единый документ. "
    "Найди: точное название поставщика (юридическое лицо из реквизитов 'Продавец' / "
    "'Грузоотправитель'), дату поставки/отгрузки, итоговую сумму С НДС. "
    "КРИТИЧЕСКИ ВАЖНО ПРО НАЗВАНИЕ ПОСТАВЩИКА: указывай ТОЛЬКО то, что физически написано на "
    "документе — ни одного слова больше. НИКОГДА не добавляй к названию слова, которые ты не "
    "видишь буквально напечатанными на фото, даже если они кажутся уместными или похожими на "
    "типичные названия компаний (например, не дописывай 'Трейд', 'Групп', 'Дистрибьюшн', 'Файн "
    "Вайнс', 'Спиритс' и подобные коммерческие слова, если их физически нет на документе). Если "
    "название на документе короткое (например, просто 'ДЖОЯ') — верни именно это короткое "
    "название, не пытайся его 'дополнить' до более полного на вид варианта. Точность важнее, "
    "чем то, что название выглядит неполным. "
    "ВСЕГДА используй кириллическое написание названия — то, что напечатано в официальном "
    "юридическом тексте реквизитов 'Продавец' / 'Грузоотправитель' (обычно начинается с "
    "'Общество с ограниченной ответственностью' или 'ИП'). Если где-то на документе (логотип, "
    "штамп, печать, фирменный знак) название стилизовано латинскими буквами (например, 'VILASH' "
    "вместо 'ВИЛАШ') — НЕ используй латинское написание, используй именно кириллическую версию "
    "из официального текста реквизитов. В таблице поставщики записаны кириллицей, латиница "
    "никогда не подойдёт для поиска совпадения. "
    "Помимо фото, тебе может быть передана ТЕКСТОВАЯ ПОДПИСЬ к сообщению — это то, что "
    "бармен вручную написал при отправке фото (например, название поставщика и/или дату). "
    "ОСОБОЕ ПРАВИЛО ДЛЯ НАЗВАНИЯ ПОСТАВЩИКА: бармен обычно точно знает, кто привёз товар, "
    "поэтому его подпись — самый надёжный источник именно для имени поставщика, надёжнее "
    "печатного текста на документе (который легко спутать из-за похожих букв при печати/сканах, "
    "например П и Т, Е и Ё, О и С). Если в подписи указано название поставщика, которое "
    "ПОХОЖЕ (созвучно, отличается на 1-2 буквы) на то, что ты прочитал на фото — считай, что "
    "это ОДНО И ТО ЖЕ название, и используй вариант из подписи бармена как итоговое значение "
    "postavshik (а не то, что 'разобрал' по фото). Например, если на фото читается 'ГРЕЙТ', а "
    "в подписи написано 'грейп' — используй 'Грейп', это одно и то же название, просто на "
    "фото буквы спутаны. "
    "Если и подпись, и фото читаются, но название поставщика в них СОВЕРШЕННО разное (не "
    "похожее, не созвучное, явно про разных поставщиков) — выбери вариант с фото как основной "
    "источник (это первичный документ), но обязательно укажи это противоречие в kommentarii и "
    "поставь uverennost 'low', чтобы менеджер проверил вручную. "
    "Для ДАТЫ применяй похожую логику: если на фото дата нечитаема — ориентируйся на дату из "
    "подписи. Если подписи нет вообще — ориентируйся только на фото, это нормальная ситуация. "
    "ПРАВИЛО ПОИСКА СУММЫ: найди итоговую строку 'Всего к оплате' / 'Итого' в самом низу "
    "товарной таблицы. В ней почти всегда напечатаны отдельно следующие три величины (читай "
    "каждую как отдельное число, не путай их между собой): "
    "(а) стоимость товаров БЕЗ налога — всего; "
    "(б) сумма налога (НДС) — отдельная величина; "
    "(в) стоимость товаров С НАЛОГОМ — всего (это и есть 'Всего к оплате' в итоге). "
    "Заполни в JSON отдельно поля summa_bez_naloga (а), summa_naloga (б) и summa (в). "
    "В поле summa должно попасть именно (в) — самое последнее (правее остальных) число в "
    "итоговой строке, в столбце с заголовком вроде 'Стоимость товаров... с налогом — всего'. "
    "НИКОГДА не вычисляй и не складывай числа сам — каждое из трёх чисел уже напечатано "
    "готовым на документе, просто прочитай и распредели их по соответствующим полям. "
    "КАНДИДАТЫ ПРИ НЕОДНОЗНАЧНОСТИ: если ты видишь НЕСКОЛЬКО правдоподобных чисел на роль "
    "итоговой суммы и не можешь точно решить, какое из них верное (например: разбивка по "
    "ставкам НДС без единой готовой суммы; или на разных страницах/фото видны разные похожие "
    "по смыслу 'итоговые' числа, и непонятно, какое из них — общий итог, а какое — частичный "
    "подытог) — НЕ выбирай и не вычисляй сам. Вместо этого заполни summa:null и заполни поле "
    "summa_candidates списком ВСЕХ правдоподобных чисел-кандидатов (просто числа, без текста), "
    "которые ты нашёл — менеджер сам выберет верное. Опиши в kommentarii коротко, что это за "
    "числа и откуда каждое взялось (например: '10% НДС = 20370.40, 22% НДС = 12546.00'). "
    "Если ты УВЕРЕН, какое число правильное — просто заполни summa этим числом, оставь "
    "summa_candidates пустым списком [], и НЕ создавай искусственную неуверенность — это "
    "должно происходить редко, только в реально сложных/неоднозначных случаях. "
    "СЛУЖЕБНЫЕ НАДПИСИ ТИПА 'Страница 1 из 3', 'Документ составлен на N листах', "
    "'Имеет продолжение на...' и подобные — это просто формальные реквизиты бланка документа, "
    "они НЕ ОЗНАЧАЮТ, что сумма на этом фото неполная или недостоверная, и НЕ ЯВЛЯЮТСЯ "
    "указанием, что тебе не хватает данных. ПОЛНОСТЬЮ ИГНОРИРУЙ такие надписи при принятии "
    "решения о сумме. "
    "Если на фото есть одна ясная итоговая строка с числом — указывай summa и ставь "
    "uverennost 'high', не создавай неуверенность искусственно. "
    "СТРОГИЙ ПОРЯДОК ПРОСМОТРА СТРАНИЦ ПРИ НЕСКОЛЬКИХ ФОТО ОДНОГО ДОКУМЕНТА: "
    "1) Сначала посмотри ИМЕННО НА ПОСЛЕДНЕЕ (по порядку получения) из переданных тебе фото. "
    "2) Найди на НЁМ строку 'Всего к оплате' / итоговую строку с числом. "
    "3) Если такая строка с числом на последнем фото есть — бери число оттуда, это и есть "
    "ответ. ИГНОРИРУЙ при этом любые похожие на вид 'итоговые' строки на ПРЕДЫДУЩИХ "
    "(более ранних) фото — даже если та строка на раннем фото выглядит полной и завершённой "
    "(с заполненными столбцами без налога / налог / с налогом) — это лишь промежуточный итог "
    "по части товаров (товарная накладная часто продолжается на следующих страницах с "
    "дополнительными позициями, и именно последняя страница содержит настоящий общий итог по "
    "всему документу). "
    "4) Только если на последнем фото вообще нет никакой итоговой строки с числами — тогда "
    "смотри на предыдущее по порядку фото, и так далее к началу. "
    "Это правило особенно важно для накладных от МИЛАРИ и похожих многостраничных документов, "
    "где первая страница часто содержит частичный, а не полный итог. "
    "Если тебе передали несколько фото ОДНОГО документа — используй итоговую строку с "
    "ПОСЛЕДНЕГО из переданных фото, где она есть. Если на разных фото видны РАЗНЫЕ по "
    "смыслу 'итоговые' числа и неясно, какое из них общий итог — это и есть случай для "
    "summa_candidates, см. выше. "
    "Работай только с тем, что реально видно на переданных тебе фото — не предполагай, не "
    "довоображай и не учитывай содержимое страниц, которых тебе не показали. "
    "Если поставщик — ООО 'МИЛАРИ', дополнительно определи покупателя по реквизитам покупателя/"
    "грузополучателя: должно быть 'Бегемот' или 'Набойщикова'. "
    "Ответь СТРОГО в виде JSON без markdown, без обратных кавычек и без пояснений, по схеме: "
    '{"postavshik":string,"pokupatel":string или "","date":string в формате ДД.MM.ГГГГ,'
    '"summa_bez_naloga":number или null,"summa_naloga":number или null,'
    '"summa":number или null,"summa_candidates":[number, ...] или [],'
    '"kommentarii":string,"uverennost":"high" или "medium" или "low"}. '
    "КРИТИЧЕСКИ ВАЖНО: твой ответ должен начинаться сразу с символа { и заканчиваться символом }. "
    "Не пиши вообще никаких слов, рассуждений или объяснений до или после JSON — даже если ты "
    "не уверен или документ неполный (например, видна только часть многостраничного документа). "
    "Любую неуверенность или замечание помещай ВНУТРЬ поля kommentarii, а не в виде текста "
    "снаружи JSON. "
    "Поле kommentarii в норме должно быть ПУСТОЙ строкой. Заполняй его только если есть "
    "существенное сомнение в дате, сумме или поставщике — тогда коротко (одна фраза) опиши "
    "именно эту неуверенность. "
    "НЕ перечисляй товарные позиции, суммы НДС по строкам или другие детали таблицы — это не нужно. "
    "Если что-то прочитать не получается уверенно — всё равно дай лучшую разумную догадку, "
    "но поставь uverennost в 'low' или 'medium'."
)


def recognize_invoice(images_base64, caption=""):
    content = [
        {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": img}}
        for img in images_base64
    ]
    user_text = "Распознай эту накладную и верни только JSON."
    if caption:
        user_text += f"\n\nТекстовая подпись бармена к этому сообщению: «{caption}»"
    content.append({"type": "text", "text": user_text})

    last_error = None
    for attempt in range(1, 3):  # одна попытка + один автоповтор
        try:
            resp = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": "claude-sonnet-4-6",
                    "max_tokens": 1000,
                    "system": SYSTEM_PROMPT,
                    "messages": [{"role": "user", "content": content}],
                },
                timeout=60,
            )
            print(f"[recognize_invoice] попытка {attempt}: HTTP {resp.status_code}, длина тела ответа: {len(resp.text)}", flush=True)
            if resp.status_code != 200:
                print(f"[recognize_invoice] тело ответа при ошибке: {resp.text[:500]}", flush=True)
            resp.raise_for_status()
            data = resp.json()
            text = "".join(block.get("text", "") for block in data.get("content", []))
            print(f"[recognize_invoice] попытка {attempt}: сырой текст ответа модели: {text!r}", flush=True)
            json_match = re.search(r"\{.*\}", text, re.DOTALL)
            clean = json_match.group(0) if json_match else text.strip()
            if not clean:
                raise ValueError(f"Пустой ответ от ИИ (попытка {attempt}). Полный ответ API: {data}")
            return json.loads(clean)
        except Exception as e:
            last_error = e
            print(f"[recognize_invoice] попытка {attempt} не удалась: {repr(e)}", flush=True)
            if attempt == 1:
                time.sleep(2)
                continue
    raise last_error


def write_to_sheet(postavshik, date, summa, pokupatel=""):
    params = {
        "sheet": SHEET_NAME,
        "supplier": postavshik,
        "date": date,
        "nomer": "",
        "summa": summa if summa is not None else "",
        "kommentarii": "",
    }
    if pokupatel:
        params["pokupatel"] = pokupatel

    r = requests.get(APPS_SCRIPT_URL, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def format_success_message(postavshik, pokupatel, date, summa, result, extra_note=""):
    pokup = f" ({pokupatel})" if pokupatel else ""
    note = f"\n📝 {extra_note}" if extra_note else ""
    return (
        f"✅ Записано: {postavshik}{pokup}\n"
        f"Дата: {date} | Сумма: {summa}"
        f"{note}\n"
        f"Строка {result.get('row')} в листе «{result.get('sheet')}»"
    )


def ask_for_sum_confirmation(parsed):
    """Отправляет менеджеру кнопки с вариантами суммы, ничего не записывая в таблицу пока."""
    conf_id = uuid.uuid4().hex[:10]
    candidates = parsed.get("summa_candidates") or []
    candidates = [c for c in candidates if c is not None]

    with confirmations_lock:
        pending_confirmations[conf_id] = {
            "postavshik": parsed.get("postavshik", ""),
            "pokupatel": parsed.get("pokupatel", ""),
            "date": parsed.get("date", ""),
            "candidates": candidates,
        }

    buttons = []
    for idx, val in enumerate(candidates):
        label = f"{val:,.2f} ₽".replace(",", " ")
        buttons.append([{"text": label, "callback_data": f"csel:{conf_id}:{idx}"}])
    buttons.append([{"text": "✏️ Другая сумма (введу сам)", "callback_data": f"cman:{conf_id}"}])

    pokup = f" ({parsed.get('pokupatel')})" if parsed.get("pokupatel") else ""
    text = (
        f"🤔 Не уверен в сумме для этой накладной.\n"
        f"Поставщик: {parsed.get('postavshik')}{pokup}\n"
        f"Дата: {parsed.get('date')}\n"
    )
    if parsed.get("kommentarii"):
        text += f"📝 {parsed.get('kommentarii')}\n"
    text += "\nВыбери верную сумму:"

    tg_send_message(MANAGER_CHAT_ID, text, reply_markup={"inline_keyboard": buttons})


def process_invoice(images_base64, caption=""):
    print(f"[process_invoice] старт, фото в накладной: {len(images_base64)}, подпись: {caption!r}", flush=True)
    try:
        parsed = recognize_invoice(images_base64, caption)
        print(f"[process_invoice] распознано: {parsed}", flush=True)
    except Exception as e:
        print(f"[process_invoice] ОШИБКА распознавания: {repr(e)}", flush=True)
        tg_send_message(MANAGER_CHAT_ID, f"⚠️ Не удалось распознать накладную.\nОшибка: {e}")
        return

    candidates = [c for c in (parsed.get("summa_candidates") or []) if c is not None]
    needs_confirmation = bool(candidates) or parsed.get("summa") is None

    if needs_confirmation:
        print(f"[process_invoice] нужна ручная проверка суммы, кандидаты: {candidates}", flush=True)
        ask_for_sum_confirmation(parsed)
        return

    try:
        result = write_to_sheet(
            parsed.get("postavshik", ""),
            parsed.get("date", ""),
            parsed.get("summa"),
            parsed.get("pokupatel", ""),
        )
        print(f"[process_invoice] результат записи: {result}", flush=True)
    except Exception as e:
        print(f"[process_invoice] ОШИБКА записи в таблицу: {repr(e)}", flush=True)
        tg_send_message(
            MANAGER_CHAT_ID,
            f"⚠️ Накладная распознана, но НЕ записалась в таблицу.\n"
            f"Ошибка: {e}\nРаспознанные данные: {parsed}",
        )
        return

    if result.get("status") == "ok":
        extra_note = ""
        if parsed.get("uverennost") in ("low", "medium"):
            extra_note = "⚠️ Не 100% уверенность в распознавании — пожалуйста, проверь эту строку."
        if parsed.get("kommentarii"):
            extra_note = (extra_note + " " if extra_note else "") + parsed.get("kommentarii")
        tg_send_message(
            MANAGER_CHAT_ID,
            format_success_message(
                parsed.get("postavshik"), parsed.get("pokupatel"), parsed.get("date"),
                parsed.get("summa"), result, extra_note,
            ),
        )
    else:
        tg_send_message(
            MANAGER_CHAT_ID,
            f"❌ Ошибка записи в таблицу: {result.get('message')}\n"
            f"Распознанные данные: {parsed}\n"
            f"Нужно занести эту накладную вручную.",
        )


def album_flusher():
    """Фоновый поток: следит за накопленными альбомами фото и обрабатывает их,
    когда новых фото в альбом не поступало некоторое время."""
    print("[album_flusher] фоновый поток запущен", flush=True)
    while True:
        try:
            time.sleep(1)
            now = time.time()
            to_process = []
            with albums_lock:
                for gid in list(pending_albums.keys()):
                    entry = pending_albums[gid]
                    if now - entry["ts"] > ALBUM_FLUSH_DELAY:
                        to_process.append({"images": entry["images"], "caption": entry["caption"]})
                        del pending_albums[gid]
            for entry_data in to_process:
                print(f"[album_flusher] обрабатываю альбом, фото: {len(entry_data['images'])}, подпись: {entry_data['caption']!r}", flush=True)
                process_invoice(entry_data["images"], entry_data["caption"])
        except Exception as e:
            print(f"[album_flusher] ОШИБКА в фоновом потоке: {repr(e)}", flush=True)


def handle_callback_query(cq):
    callback_id = cq["id"]
    data = cq.get("data", "")
    chat_id = cq["message"]["chat"]["id"]
    message_id = cq["message"]["message_id"]

    tg_answer_callback_query(callback_id)
    tg_remove_keyboard(chat_id, message_id)

    if data.startswith("csel:"):
        _, conf_id, idx_str = data.split(":")
        idx = int(idx_str)
        with confirmations_lock:
            entry = pending_confirmations.pop(conf_id, None)
        if not entry:
            tg_send_message(chat_id, "⚠️ Эта накладная уже обработана или устарела.")
            return
        try:
            chosen_summa = entry["candidates"][idx]
        except IndexError:
            tg_send_message(chat_id, "⚠️ Не удалось определить выбранный вариант.")
            return
        finalize_confirmed_invoice(chat_id, entry, chosen_summa)

    elif data.startswith("cman:"):
        _, conf_id = data.split(":")
        with confirmations_lock:
            entry = pending_confirmations.get(conf_id)
        if not entry:
            tg_send_message(chat_id, "⚠️ Эта накладная уже обработана или устарела.")
            return
        with awaiting_lock:
            awaiting_manual_sum[chat_id] = conf_id
        tg_send_message(chat_id, "Напиши правильную сумму одним сообщением (например: 12345.67)")


def finalize_confirmed_invoice(chat_id, entry, chosen_summa):
    try:
        result = write_to_sheet(entry["postavshik"], entry["date"], chosen_summa, entry["pokupatel"])
        print(f"[finalize_confirmed_invoice] результат записи: {result}", flush=True)
    except Exception as e:
        tg_send_message(chat_id, f"⚠️ Не удалось записать в таблицу: {e}")
        return

    if result.get("status") == "ok":
        tg_send_message(
            chat_id,
            format_success_message(entry["postavshik"], entry["pokupatel"], entry["date"], chosen_summa, result),
        )
    else:
        tg_send_message(chat_id, f"❌ Ошибка записи в таблицу: {result.get('message')}")


SKIP_SUPPLIER_KEYWORDS = ["кухня", "рынок", "магазин"]


def parse_sverka_file(file_bytes):
    """Разбирает xlsx-файл сверки от бухгалтера на блоки по поставщикам.
    Возвращает список {"name": str, "invoices": [{"date": "ДД.MM.ГГГГ", "summa": float, "concept": str}]}
    Пропускает внутренние псевдо-поставщики (кухня/рынок/магазин)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    suppliers = []
    current_name = None
    current_invoices = []

    def flush():
        if current_name and current_invoices:
            name_lower = current_name.lower()
            if not any(k in name_lower for k in SKIP_SUPPLIER_KEYWORDS):
                suppliers.append({"name": current_name, "invoices": current_invoices})

    for row in ws.iter_rows(min_row=5):
        b_val = row[1].value if len(row) > 1 else None
        if isinstance(b_val, str) and b_val.strip().startswith("Поставщик/Покупатель"):
            flush()
            name = b_val.split(":", 1)[1].strip() if ":" in b_val else b_val.strip()
            name = name.lstrip(":").strip()
            current_name = name
            current_invoices = []
            continue

        c_val = row[2].value if len(row) > 2 else None
        if c_val:
            d_val = row[3].value if len(row) > 3 else None
            g_val = row[6].value if len(row) > 6 else None
            k_val = row[10].value if len(row) > 10 else ""
            if d_val is not None and g_val is not None:
                if hasattr(d_val, "strftime"):
                    date_str = d_val.strftime("%d.%m.%Y")
                else:
                    date_str = str(d_val).split(" ")[0]
                try:
                    summa_val = round(float(g_val), 2)
                except (TypeError, ValueError):
                    continue
                current_invoices.append({
                    "date": date_str,
                    "summa": summa_val,
                    "concept": str(k_val) if k_val else "",
                })

    flush()
    return suppliers


def run_sverka(file_bytes):
    suppliers = parse_sverka_file(file_bytes)
    print(f"[run_sverka] поставщиков для сверки: {len(suppliers)}", flush=True)

    resp = requests.post(
        APPS_SCRIPT_URL,
        json={"action": "compare", "sheet": SHEET_NAME, "suppliers": suppliers},
        timeout=120,
    )
    resp.raise_for_status()
    return resp.json()


def format_sverka_summary(result):
    if result.get("status") != "ok":
        return f"❌ Ошибка сверки: {result.get('message')}"

    results = result.get("results", [])
    total_matched = sum(r.get("matched", 0) for r in results if "error" not in r)
    total_mismatches = sum(len(r.get("mismatches", [])) for r in results if "error" not in r)
    total_missing_in_buh = sum(len(r.get("missingInBuh", [])) for r in results if "error" not in r)
    errors = [r for r in results if "error" in r]

    lines = [
        "📊 Сверка завершена",
        f"✅ Совпало: {total_matched} накладных",
    ]
    if total_mismatches or total_missing_in_buh:
        lines.append(f"⚠️ Расхождений: {total_mismatches + total_missing_in_buh}")
    lines.append("")

    for r in results:
        if "error" in r:
            continue
        mism = r.get("mismatches", [])
        missing = r.get("missingInBuh", [])
        if not mism and not missing:
            continue
        lines.append(f"— {r['supplier']} —")
        for m in mism:
            if m["type"] == "сумма не совпадает":
                lines.append(f"  {m['date']}: у буха {m['buhSumma']}, у нас {m['ourSumma']}")
            else:
                lines.append(f"  {m['date']}: {m['buhSumma']} — нет в нашем реестре")
        for mb in missing:
            lines.append(f"  {mb['date']}: {mb['summa']} — есть у нас, нет у буха")
        lines.append("")

    if errors:
        lines.append("⚠️ Не удалось сверить:")
        for e in errors:
            lines.append(f"  {e['supplier']}: {e['error']}")

    return "\n".join(lines)


def send_long_message(chat_id, text):
    """Telegram режет сообщения длиннее ~4096 символов — разбиваем по частям."""
    LIMIT = 3800
    while text:
        chunk = text[:LIMIT]
        tg_send_message(chat_id, chunk)
        text = text[LIMIT:]


@app.route("/webhook", methods=["POST"])
def webhook():
    ensure_flusher_started()
    update = request.get_json(silent=True) or {}
    print(f"[webhook] получено обновление: {update}", flush=True)

    if "callback_query" in update:
        handle_callback_query(update["callback_query"])
        return "ok"

    msg = update.get("message")

    if not msg:
        print("[webhook] в обновлении нет message, пропускаю", flush=True)
        return "ok"

    chat_id = msg["chat"]["id"]

    if "text" in msg:
        text = msg["text"].strip()
        if text == "/start":
            tg_send_message(
                chat_id,
                "Привет! Присылай сюда фото накладных — я распознаю и сам занесу в реестр. "
                "Если накладная на несколько страниц — отправляй фото как альбом (выбери все сразу).",
            )
            return "ok"

        with awaiting_lock:
            conf_id = awaiting_manual_sum.pop(chat_id, None)
        if conf_id:
            try:
                manual_summa = float(text.replace(",", ".").replace(" ", ""))
            except ValueError:
                tg_send_message(chat_id, "Не получилось понять число. Напиши просто сумму, например: 12345.67")
                with awaiting_lock:
                    awaiting_manual_sum[chat_id] = conf_id
                return "ok"
            with confirmations_lock:
                entry = pending_confirmations.pop(conf_id, None)
            if not entry:
                tg_send_message(chat_id, "⚠️ Эта накладная уже обработана или устарела.")
                return "ok"
            finalize_confirmed_invoice(chat_id, entry, manual_summa)
            return "ok"

    if "photo" not in msg:
        if "document" in msg:
            doc = msg["document"]
            file_name = doc.get("file_name", "")
            if file_name.lower().endswith((".xlsx", ".xls")):
                tg_send_message(chat_id, "🔄 Получил файл сверки, сравниваю с реестром...")
                try:
                    file_path = tg_get_file_path(doc["file_id"])
                    file_bytes = tg_download_file(file_path)
                    result = run_sverka(file_bytes)
                    summary = format_sverka_summary(result)
                    send_long_message(chat_id, summary)
                except Exception as e:
                    print(f"[webhook] ОШИБКА сверки: {repr(e)}", flush=True)
                    tg_send_message(chat_id, f"⚠️ Не удалось выполнить сверку: {e}")
            else:
                tg_send_message(chat_id, "Это не похоже на xlsx-файл сверки — присылай файл от бухгалтера в формате Excel.")
        else:
            print("[webhook] в message нет photo и нет document, пропускаю", flush=True)
        return "ok"

    file_id = msg["photo"][-1]["file_id"]  # самое крупное по размеру фото
    media_group_id = msg.get("media_group_id")
    caption = (msg.get("caption") or "").strip()
    print(f"[webhook] фото получено, media_group_id={media_group_id}, caption={caption!r}", flush=True)

    try:
        file_path = tg_get_file_path(file_id)
        img_bytes = tg_download_file(file_path)
        img_b64 = base64.b64encode(img_bytes).decode()
        print(f"[webhook] фото скачано, размер base64: {len(img_b64)}", flush=True)
    except Exception as e:
        print(f"[webhook] ОШИБКА скачивания фото: {repr(e)}", flush=True)
        tg_send_message(MANAGER_CHAT_ID, f"⚠️ Не удалось загрузить фото из Telegram: {e}")
        return "ok"

    if media_group_id:
        with albums_lock:
            entry = pending_albums.setdefault(media_group_id, {"images": [], "caption": "", "ts": time.time()})
            entry["images"].append(img_b64)
            if caption:
                entry["caption"] = caption
            entry["ts"] = time.time()
        print(f"[webhook] фото добавлено в альбом {media_group_id}, всего в альбоме: {len(entry['images'])}", flush=True)
    else:
        print("[webhook] одиночное фото, обрабатываю сразу (синхронно)", flush=True)
        process_invoice([img_b64], caption)

    return "ok"


@app.route("/", methods=["GET"])
def health():
    return "Бот работает"


_flusher_started = False
_flusher_lock = threading.Lock()


def ensure_flusher_started():
    global _flusher_started
    if _flusher_started:
        return
    with _flusher_lock:
        if not _flusher_started:
            threading.Thread(target=album_flusher, daemon=True).start()
            _flusher_started = True
            print("[ensure_flusher_started] фоновый поток запущен внутри воркера", flush=True)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
